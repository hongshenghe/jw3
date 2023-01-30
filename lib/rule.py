# -*- encoding: utf-8 -*-
'''
@File    :   rule.py
@Time    :   2023/01/07 22:05:27
@Author  :   Hongsheng He
@Version :   1.0
@Contact :   24836227@qq.com
@License :   (C)Copyright 2007-2023, hongsheng
@Desc    :   规则管理
'''


import os

import openpyxl
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from lib.base import ensurePath, generate_batchid
from lib.logger import logging
from lib.utils._05 import (GenerateProjectSiteInfo, GetProjectDict, SetNone,
                           SetValue, GetProjectSite, GetRackProductLine, GetPosition, GetAssetInfo, GetSNMPVersion, GetRackHeight)
from lib.utils._06 import ResourceType, DeviceCode, DeviceIP, VlanAsset
from lib.utils._09 import Copy, GetDict, GetSubProductLine, GetNetworkLogicCode
from lib.utils._12 import GetDataCenterShort, BMC, HTTPS
from lib.utils._13 import GetNetwork4AWebAssetName
from lib.utils._15 import PrometheusSNMPServer, PrometheusFilter, GetPrometheusAssetInfo
from lib.utils._51 import SequenceNumber, SnmpTarget
from lib.utils._52 import GetVMHostInfo, GetVMInfo
from lib.utils._62 import GetCloudDesktopRack, GetCloudDesktopPos, GetCloudDesktopProjectSite
from lib.utils._16 import GetMaintenanceInfo, GetShortSiteName, GetMaintenanceColumn
from lib.zero import JWZero
from lib.dict import JWDict


class JWRule(object):

    zero = None
    jwDict = None

    # 输入文件编号
    file_id = ""

    # 输出文件编号
    output_file_id = ""

    # 文件名称
    file_name = ""
    work_dir = ""
    batch_id = ""

    # 配置文件原始内容
    content = ""
    # 配置文件名称
    name = ""
    # 配置文件meta信息
    meta = []
    # 配置文件规则
    rules = []

    error = None

    def __init__(self, batch_id: str, file_id: str, work_dir: str, zero: JWZero, jwDict: JWDict) -> None:

        self.file_name = "%s.yaml" % (file_id)
        full_name = os.path.join(work_dir, "rules",  self.file_name)

        logging.info("full_name: %s" % full_name)

        if not os.path.exists(full_name):
            return

        self.work_dir = work_dir
        self.batch_id = batch_id
        self.zero = zero
        self.jwDict = jwDict
        self.file_id = file_id

    def Generate(self) -> Exception:
        # 生成路径
        output_dir = os.path.join(self.work_dir, "down", self.batch_id)
        ensurePath(output_dir)

        # load
        err = self.load()
        if err:
            return err

        # generate data

        # 数据分类
        dynamic_rule_list, fixed_rule_list, copy_rule_list = [], [], []
        for sheet in self.rules:
            sheet_name = sheet['sheet_name']
            sheet_type = "动态计算" if "sheet_type" not in sheet else sheet['sheet_type']
            if sheet_type == "动态计算":
                dynamic_rule_list.append(sheet)
            if sheet_type == "fixed":
                fixed_rule_list.append(sheet)
            if sheet_type == "copy":
                copy_rule_list.append(sheet)

        # 动态数据计算
        file_name = self.generate_output_file_name()
        output = os.path.join(output_dir, '%s.xlsx' % file_name)

        data = {}
        for sheet in dynamic_rule_list:

            sheet_name = sheet['sheet_name']
            sheet_type = "动态计算" if "sheet_type" not in sheet else sheet['sheet_type']
            order = None if "order" not in sheet else sheet['order']
            content = None if "content" not in sheet else sheet['content']

            logging.info("处理 %s-%s 开始" % (self.file_id, sheet_name))
            logging.info("  sheet类型: %s" % (sheet_type))

            df = pd.DataFrame()
            col_id = 1
            for r in sheet['columns']:
                col_name = r['name']
                value = None if "value" not in r else r['value']
                source_sheet = None if "source_sheet" not in r else r['source_sheet']
                source_column = None if "source_column" not in r else r['source_column']

                logging.info("处理 %s-%s，第%s列,%s" %
                             (self.file_id, sheet_name, col_id, col_name))
                df, flag = eval(r['method'])(
                    self.zero, self.jwDict, df, col_name, value, source_sheet, source_column)

                col_id += 1
            data[sheet_name] = df
            logging.info("处理 %s-%s 结束" % (self.file_id, sheet_name))

        if len(data) != 0:
            self._save_excel(output, data)

        # 固定表格生成
        for sheet in fixed_rule_list:
            sheet_name = sheet['sheet_name']
            sheet_type = "动态计算" if "sheet_type" not in sheet else sheet['sheet_type']
            order = None if "order" not in sheet else sheet['order']
            # width = 50 if "width" not in sheet else sheet['width']
            content = None if "content" not in sheet else sheet['content']

            logging.info("处理 %s-%s 开始" % (self.file_id, sheet_name))
            logging.info("  sheet类型: %s" % (sheet_type))
            logging.info("  位置: %s" % (order))

            if sheet_type == "fixed":
                self._save_fixed_sheet(
                    output, sheet_name, order, content)
            logging.info("处理 %s-%s 结束" % (self.file_id, sheet_name))

        # 复制sheet
        for sheet in copy_rule_list:
            source_file = self.zero.Full_file_name()
            source_sheet = None if "source_sheet" not in sheet else sheet['source_sheet']

            target_file = output
            target_sheet_name = sheet['sheet_name']

            err = self._copy_sheet(
                source_file, source_sheet, target_file, target_sheet_name)

            if err:
                return err

        # 增加作者信息 output
        workbook = openpyxl.load_workbook(output)
        workbook.properties.creator = "理想"
        workbook.properties.description = file_name
        workbook.save(output)

        return None

    def _copy_sheet(self, source_file, source_sheet, target_file, target_sheet_name) -> Exception:
        """复制sheet

        Args:
            source_file (_type_): 源数据文件
            source_sheet (_type_): 源sheet
            target_file (_type_): 目标文件
            target_sheet_name (_type_): 目标sheet名称
        """
        # print("_copy_sheet  source_file：%s, source_sheet：%s, target_file：%s, target_sheet_name：%s" %
        #       source_file, source_sheet, target_file, target_sheet_name)
        # print("_copy_sheet  source_file：%s" % source_file)
        # print("_copy_sheet  source_sheet：%s" % source_sheet)
        # print("_copy_sheet  target_file：%s" % target_file)
        # print("_copy_sheet  target_sheet_name：%s" % target_sheet_name)
        if not os.path.exists(source_file):
            err = "%s文件名不存在，无法复制sheet,%s-%s" % (source_file,
                                                target_file, target_sheet_name)
            logging.error(err)
            return Exception(err)

        wb = openpyxl.load_workbook(source_file)
        sheets = wb.sheetnames
        # print("sheets:%s" % sheets)
        if source_sheet not in sheets:
            err = "%s文件名%s sheet不存在，无法复制sheet,%s-%s" % (source_file, source_sheet,
                                                        target_file, target_sheet_name)
            logging.error(err)
            return Exception(err)

        work_sheet = wb[source_sheet]

        if os.path.exists(target_file):
            os.remove(target_file)

        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = target_sheet_name

        for row in work_sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # new_workbook.properties.author = "理想"
        new_workbook.save(target_file)

        return None

    def _save_fixed_sheet(self, output_file, sheet_name, order,   content):

        if not os.path.exists(output_file):
            logging.error("%s 文件不存在，不能增加新sheet" % output_file)
            return

        wb = openpyxl.load_workbook(output_file)
        if order == "first":
            ws = wb.create_sheet(title=sheet_name, index=0)
        else:
            ws = wb.create_sheet(title=sheet_name)
        # ws.sheet_properties.tabColor = 'ff72BA'

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        row_num = 1
        for c in content:
            col_num = 1
            for value in c.split("|"):
                cell = ws.cell(column=col_num, row=row_num)
                cell.value = "\n".join(value.replace("`", "|").split(" "))

                cell.border = border
                cell.alignment = Alignment(wrapText=True)

                col_num += 1

            row_num += 1

        # print("sheet_name: %s" % ws.title)
        ws.column_dimensions["A"].column_width = 50
        if "说明" in ws.title:
            ws.column_dimensions["A"].width = 10.27
            ws.column_dimensions["B"].width = 77.45
        if "字典" in ws.title:
            # print("字典命中：", ws.title)
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 60

        ws.column_dimensions["B"].width = 50

        # ws.column_dimensions["A"].column_width = None
        # ws.column_dimensions["B"].column_width = None
        # ws.column_dimensions["A"].auto_size = True
        # ws.column_dimensions["B"].auto_size = True
        # ws.column_dimensions["A"].bestFit = True
        # ws.column_dimensions["B"].bestFit = True
        # wb.properties.author = "理想"
        wb.save(output_file)

    def _save_excel(self, file_name: str, data: dict) -> None:
        """保存文件"""

        writer = pd.ExcelWriter(file_name)
        for k, v in data.items():
            sheet_name = k
            df = v

            df.to_excel(writer, sheet_name=sheet_name,
                        index=False, engine='openpyxl')

        writer.close()

    def load(self) -> Exception:
        # 加载配置文件
        fn = os.path.join(self.work_dir, "rules", self.file_name)
        # print("fn:", fn)
        if not os.path.exists(fn):
            err = Exception("配置文件不存在:%s" % fn)
            self.error = err
            return err

        # logging.info("fn:",fn)
        content = ''
        with open(fn, 'r', encoding="utf-8") as f:
            content = yaml.full_load(f)

        self.content = content
        self.rules = self.content['rules']
        self.meta = self.content['meta']
        self.name = self.content['meta']['name']
        self.output_file_id = str(self.content['meta']['fileID'])

        return None

    def generate_output_file_name(self) -> str:
        file_no = self.output_file_id.zfill(2)
        project_name = self.zero.GetProject("项目名称")
        return "%s%s--%s" % (file_no,
                             project_name, self.name)


# 加载所有规则
def LoadRules(work_dir: str, zero: JWZero, jwDict: JWDict) -> list:
    logging.info("加载所有规则文件")
    rules = []

    rule_dir = os.path.join(work_dir, "rules")
    if not os.path.exists(rule_dir):
        return rules

    batch_id = generate_batchid()

    rule_file_list = [f for f in os.listdir(
        rule_dir) if os.path.splitext(f)[1] == ".yaml"]

    rule_file_id_list = []
    for f in rule_file_list:
        file_id = str(os.path.splitext(f)[0])
        rule_file_id_list.append(file_id)

    rule_file_id_list.sort()
    for file_id in rule_file_id_list:
        logging.info("加载:%s.yaml" % file_id)
        rules.append(JWRule(batch_id, file_id, work_dir, zero, jwDict))

    return rules
